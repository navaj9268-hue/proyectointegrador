import openpyxl
from openpyxl.styles import Font, PatternFill
from faker import Faker
import random
from datetime import datetime, timedelta
import os

# Inicializar Faker con locale español
fake = Faker('es_ES')

# Crear workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remover hoja por defecto

# Definir estilos
header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
header_font = Font(bold=True)

def format_header(sheet):
    """Formatear la fila de encabezados"""
    for cell in sheet[1]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
    
    # Auto-ajustar ancho de columnas
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width

# TABLA HOTELS
ws_hotels = wb.create_sheet("Hotels")
headers = ["ID", "Nombre", "Dirección", "Teléfono", "Email", "Descripción", "Creado", "Actualizado"]
ws_hotels.append(headers)

for i in range(1, 101):
    ws_hotels.append([
        i,
        fake.company(),
        fake.address(),
        fake.phone_number(),
        fake.email(),
        fake.sentence(),
        fake.date_time_this_year(),
        fake.date_time_this_year()
    ])
format_header(ws_hotels)

# TABLA GUESTS
ws_guests = wb.create_sheet("Guests")
headers = ["ID", "Nombre", "Email", "Teléfono", "Número de Documento", "Creado", "Actualizado"]
ws_guests.append(headers)

for i in range(1, 101):
    ws_guests.append([
        i,
        fake.name(),
        fake.email(),
        fake.phone_number(),
        fake.numerify(text='###########'),
        fake.date_time_this_year(),
        fake.date_time_this_year()
    ])
format_header(ws_guests)

# TABLA ROOMS
ws_rooms = wb.create_sheet("Rooms")
headers = ["ID", "Hotel ID", "Número", "Tipo", "Precio", "Estado", "Notas", "Creado", "Actualizado"]
ws_rooms.append(headers)

room_types = ['Single', 'Double', 'Suite', 'Deluxe', 'Penthouse']
room_statuses = ['available', 'occupied', 'maintenance']

for i in range(1, 101):
    ws_rooms.append([
        i,
        random.randint(1, 100),
        f"{i}A",
        random.choice(room_types),
        random.randint(50, 500),
        random.choice(room_statuses),
        fake.sentence(),
        fake.date_time_this_year(),
        fake.date_time_this_year()
    ])
format_header(ws_rooms)

# TABLA RESERVATIONS
ws_reservations = wb.create_sheet("Reservations")
headers = ["ID", "Hotel ID", "Room ID", "Guest ID", "Check-in", "Check-out", "Total", "Estado", "Notas", "Creado", "Actualizado"]
ws_reservations.append(headers)

reservation_statuses = ['booked', 'checked_in', 'checked_out', 'cancelled']

for i in range(1, 101):
    checkin_date = fake.date_object()
    checkout_date = checkin_date + timedelta(days=random.randint(1, 7))
    
    ws_reservations.append([
        i,
        random.randint(1, 100),
        random.randint(1, 100),
        random.randint(1, 100),
        checkin_date,
        checkout_date,
        random.randint(100, 1000),
        random.choice(reservation_statuses),
        fake.sentence(),
        fake.date_time_this_year(),
        fake.date_time_this_year()
    ])
format_header(ws_reservations)

# TABLA PAYMENTS
ws_payments = wb.create_sheet("Payments")
headers = ["ID", "Reservation ID", "Monto", "Método", "ID Transacción", "Nombre Pagador", "Notas", "User ID", "Creado", "Actualizado"]
ws_payments.append(headers)

payment_methods = ['efectivo', 'tarjeta', 'transferencia', 'cheque']

for i in range(1, 101):
    ws_payments.append([
        i,
        random.randint(1, 100),
        random.randint(50, 1000),
        random.choice(payment_methods),
        fake.uuid4(),
        fake.name(),
        fake.sentence(),
        random.randint(1, 100),
        fake.date_time_this_year(),
        fake.date_time_this_year()
    ])
format_header(ws_payments)

# TABLA BOOKINGS
ws_bookings = wb.create_sheet("Bookings")
headers = ["ID", "User ID", "Room ID", "Check-in", "Check-out", "Estado", "Total", "Creado", "Actualizado"]
ws_bookings.append(headers)

booking_statuses = ['reserved', 'checked_in', 'completed', 'cancelled']

for i in range(1, 101):
    checkin_date = fake.date_object()
    checkout_date = checkin_date + timedelta(days=random.randint(1, 7))
    
    ws_bookings.append([
        i,
        random.randint(1, 100),
        random.randint(1, 100),
        checkin_date,
        checkout_date,
        random.choice(booking_statuses),
        random.randint(100, 1000),
        fake.date_time_this_year(),
        fake.date_time_this_year()
    ])
format_header(ws_bookings)

# TABLA INVENTORIES
ws_inventories = wb.create_sheet("Inventories")
headers = ["ID", "Hotel ID", "Artículo", "Cantidad", "Ubicación", "Notas", "Creado", "Actualizado"]
ws_inventories.append(headers)

inventory_items = ['Almohadas', 'Sábanas', 'Toallas', 'Vasos', 'Platos', 'Cucharas', 'Cuchillos', 'Tenedores', 'Lámparas', 'Cortinas']

for i in range(1, 101):
    ws_inventories.append([
        i,
        random.randint(1, 100),
        random.choice(inventory_items),
        random.randint(10, 500),
        f"Piso {random.randint(1, 10)}, Sala {random.randint(1, 5)}",
        fake.sentence(),
        fake.date_time_this_year(),
        fake.date_time_this_year()
    ])
format_header(ws_inventories)

# TABLA VEHICLES
ws_vehicles = wb.create_sheet("Vehicles")
headers = ["ID", "Reservation ID", "Placa", "Marca", "Modelo", "Color", "Estado", "Fecha Entrada", "Fecha Salida", "Lugar Estacionamiento", "Notas", "Creado", "Actualizado"]
ws_vehicles.append(headers)

vehicle_statuses = ['parking', 'left']
brands = ['Toyota', 'Honda', 'Ford', 'Chevrolet', 'BMW', 'Mercedes', 'Audi', 'Nissan', 'Hyundai', 'Kia']
colors = ['Blanco', 'Negro', 'Plata', 'Rojo', 'Azul', 'Verde', 'Gris', 'Oro']

for i in range(1, 101):
    entry_date = fake.date_time_this_year()
    exit_date = entry_date + timedelta(days=random.randint(1, 7))
    
    ws_vehicles.append([
        i,
        random.randint(1, 100),
        fake.license_plate(),
        random.choice(brands),
        fake.word(),
        random.choice(colors),
        random.choice(vehicle_statuses),
        entry_date,
        exit_date,
        f"Piso {random.randint(1, 5)}, Espacio {random.randint(1, 50)}",
        fake.sentence(),
        fake.date_time_this_year(),
        fake.date_time_this_year()
    ])
format_header(ws_vehicles)

# Guardar archivo en Descargas
downloads_folder = os.path.expanduser('~/Downloads')
filename = f"database_sample_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
filepath = os.path.join(downloads_folder, filename)
wb.save(filepath)
print(f"✓ Archivo Excel creado exitosamente: {filepath}")
